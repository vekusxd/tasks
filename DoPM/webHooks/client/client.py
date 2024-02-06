from flask import Flask, request
from datetime import datetime
import os.path
import openpyxl

app = Flask(__name__)


@app.route("/", methods=["POST"])
def index():
    # with open("counter.txt", "r") as f:
    #     content = f.readline()
    
    # counter = int(content)

    # print("Counter {number}".format(number = counter))

    # if os.path.exists("purchases.xlsx"):
    #     book = openpyxl.Workbook()
    #     print("File exists")
    # else:
    #     book = openpyxl.Workbook()
    #     book.save("purchases.xlsx")
    #     book = openpyxl.Workbook()

    # sheet = book.active
    
    # value = request.json
    # userId = value["_id"]
    # currentTime = datetime.now().strftime("%d/%m/%Y, %H:%M:%S")
    # print(currentTime)
    # print(userId)
    # for item in value["goods"]:
    #     # sheet.append(counter)
    #     # sheet.append(userId)
    #     # sheet.append(currentTime)
    #     # sheet.append(item["item"])
    #     # sheet.append(item["price"])
    #     # sheet["C{0}".format(counter)] = currentTime
    #     # sheet["D{0}".format(counter)] = item["item"]
    #     # sheet["E{0}".format(counter)] = item["price"]
    #     # sheet["E{0}".format(counter)].append(counter)
    #     # sheet["E{0}".format(counter)].append(counter)
    #     sheet["A{0}".format(counter)].append(counter)
    #     sheet["B{0}".format(counter)].append(userId)
    #     sheet["C{0}".format(counter)].append(currentTime)
    #     sheet["D{0}".format(counter)].append(item["item"])
    #     sheet["E{0}".format(counter)].append(item["price"])
    #     book.save("purchases.xlsx")
    #     counter += 1
    
    # book.close()

    # with open("counter.txt", "w") as f:
    #     f.write(str(counter))

    counter = 0

    with open("counter.txt", "r") as f:
        content = f.readline()
        counter = int(content)
        f.close()
    

    if os.path.exists("purchases.xlsx"):
        print("File exists")
    else:
        book = openpyxl.Workbook()
        book.save("purchases.xlsx")
        book.close()
    
    book = openpyxl.load_workbook("purchases.xlsx")
    
    sheet = book.active

    value = request.json

    userId = value["_id"]
    currentTime = datetime.now().strftime("%d/%m/%Y, %H:%M:%S")

    
    for item in value["goods"]:
        sheet["A{0}".format(counter)] = counter
        sheet["B{0}".format(counter)] = userId
        sheet["C{0}".format(counter)] = currentTime
        sheet["D{0}".format(counter)] = item["item"]
        sheet["E{0}".format(counter)] = item["price"]
        counter += 1

    book.save("purchases.xlsx")
    book.close()


    with open("counter.txt", "w") as f:
        f.write(str(counter))
        f.close()

    return "OK"

if __name__=="__main__":
    app.run()